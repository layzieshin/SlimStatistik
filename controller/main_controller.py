import configparser
import datetime as dt
import os
from typing import List, Tuple, Dict, Optional

from models.repository import Repository


class MainController:
    """
    Application-LOGIC (Use-Cases). Kein UI, kein SQL – nur Domänenlogik.
    """
    def __init__(self, settings_path: Optional[str] = None, mapping_path: Optional[str] = None):
        self.settings_path = settings_path or os.path.join("config", "settings.ini")
        os.makedirs(os.path.dirname(self.settings_path), exist_ok=True)

        cfg = configparser.ConfigParser()
        if os.path.exists(self.settings_path):
            cfg.read(self.settings_path, encoding="utf-8")

        # Default-Pfade
        self.paths: Dict[str, str] = {
            "database_path": "",
            "excel_file": os.path.join("export", "DeletedSuspects.xlsx"),
            "export_dir": "export",
        }
        if "paths" in cfg:
            self.paths.update(cfg["paths"])

        # Analyten-Filter (aus Suche ausschließen)
        self._excluded = set()
        if "filters" in cfg:
            ex = cfg["filters"].get("exclude_analytes", "")
            self._excluded = set([x.strip() for x in ex.split(";") if x.strip()])

        self._mapping_path = mapping_path
        self.repo = Repository(self.paths.get("database_path", ""))

    # ---------------------- Settings
    def save_settings(self):
        cfg = configparser.ConfigParser()
        cfg["paths"] = dict(self.paths)
        cfg["filters"] = {"exclude_analytes": ";".join(sorted(self._excluded))}
        os.makedirs(os.path.dirname(self.settings_path), exist_ok=True)
        with open(self.settings_path, "w", encoding="utf-8") as f:
            cfg.write(f)

    # ---------------------- Analyten-Listen
    def list_all_analytes(self) -> List[str]:
        try:
            return self.repo.list_all_analytes()
        except Exception:
            return []

    def get_excluded_analytes(self) -> List[str]:
        return sorted(self._excluded)

    def update_excluded_analytes(self, items: List[str]) -> None:
        self._excluded = set(items)
        self.save_settings()

    def list_included_analytes(self) -> List[str]:
        all_codes = set(self.list_all_analytes())
        return sorted([a for a in all_codes if a not in self._excluded])

    # ---------------------- Monats-Export (Hook – aktuell als Platzhalter)
    def monthly_export_if_first(self) -> None:
        return

    # ---------------------- Zählungen
    def build_counts_rows_multi(
        self,
        start: dt.datetime,
        end: dt.datetime,
        analytes: List[str],
        only_open_weekdays: bool,
    ) -> List[Tuple[str, str, str, str]]:
        s = start.strftime("%Y-%m-%d %H:%M:%S")
        e = end.strftime("%Y-%m-%d %H:%M:%S")

        rows: List[Tuple[str, str, str, str]] = []
        if analytes:
            for code, cnt in self.repo.count_requirements_per_analyte(analytes, s, e):
                rows.append((f"Anforderungen {code}", str(cnt), "", ""))

        total, open_cnt, done_cnt = self.repo.count_befund_status(s, e)
        rows += [
            ("Befunde (offen)",  str(open_cnt), "", ""),
            ("Befunde (fertig)", str(done_cnt), "", ""),
            ("Befunde (alle)",   str(total),    "", ""),
        ]

        wd = self.repo.count_befunde_per_weekday(s, e, only_open=only_open_weekdays, analytes=None)
        avg = f"{(sum(wd.values())/7.0):.2f}" if wd else "0.00"
        rows.append((f"Wochentage ({'nur offene' if only_open_weekdays else 'alle'})", "", "Durchschnitt", avg))
        for day in ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]:
            rows.append((f"  {day}", str(wd.get(day, 0)), "", ""))

        return rows

    # ---------------------- Offene Anforderungen
    def build_open_counts_since(self, analytes: List[str], since: dt.datetime) -> List[Tuple[str, int]]:
        s = since.strftime("%Y-%m-%d %H:%M:%S")
        return self.repo.count_open_requirements_per_analyte(analytes, s)

    # ---------------------- Nicht entnommen?
    def suspected_missing_blood_draw(self) -> List[Dict]:
        return self.repo.list_suspected_missing_draw(older_than_hours=24)

    # === Excel-Audit Helpers ==================================================
    @staticmethod
    def _ensure_audit_ws(wb, desired_headers: List[str], sheet_name: str = "Deleted Suspects"):
        """
        Sorgt dafür, dass es ein Worksheet mit Header **in Zeile 1** gibt
        (ohne doppelte Header und ohne append() auf „leere Zeile 1“).
        """
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active if wb.active else wb.create_sheet(title=sheet_name)
            ws.title = sheet_name

        # Aktuelle Header aus Zeile 1 lesen
        current = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        current = [str(h) for h in current if h is not None and str(h) != ""]

        if not current:
            # Neue Header EXPLIZIT in Row 1 setzen (kein append!)
            for idx, name in enumerate(desired_headers, start=1):
                ws.cell(row=1, column=idx, value=str(name))
            current = desired_headers[:]
        else:
            # Upgrade: fehlende Spalten hinten anhängen (nur Row1 beschreiben)
            for h in desired_headers:
                if h not in current:
                    current.append(h)
            for idx, name in enumerate(current, start=1):
                ws.cell(row=1, column=idx, value=str(name))

        return ws, current

    @staticmethod
    def _ensure_table(ws, headers: List[str], table_name: str = "DeletedSuspectsTable"):
        """
        Erstellt/aktualisiert eine Excel-Tabelle über dem Datenbereich.
        Nur, wenn es mind. 1 Datenzeile gibt (>=2 Zeilen inkl. Header),
        sonst wird KEINE Tabelle angelegt (vermeidet Excel-Reparatur).
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        last_row = ws.max_row
        last_col = len(headers)
        if last_row < 2:
            return

        ref = f"A1:{get_column_letter(last_col)}{last_row}"

        if table_name in ws.tables:
            t = ws.tables[table_name]
            t.ref = ref
        else:
            t = Table(displayName=table_name, ref=ref)
            style = TableStyleInfo(name="TableStyleMedium2",
                                   showFirstColumn=False,
                                   showLastColumn=False,
                                   showRowStripes=True,
                                   showColumnStripes=False)
            t.tableStyleInfo = style
            ws.add_table(t)

    def _append_audit_rows(self, infos: List[Dict]):
        """Appendet mehrere Audit-Zeilen in EINEM Excel-Open/Save – ohne doppelte Header."""
        try:
            from openpyxl import Workbook, load_workbook
        except Exception:
            return

        os.makedirs(os.path.dirname(self.paths["excel_file"]), exist_ok=True)
        desired_headers = [
            "Zeitpunkt", "ProbenNr", "Name", "VName",
            "GebDat", "PatID",
            "Abnahme", "AuftragsNr", "Einsender", "Analyte"
        ]

        if os.path.exists(self.paths["excel_file"]):
            wb = load_workbook(self.paths["excel_file"])
        else:
            wb = Workbook()

        ws, headers = self._ensure_audit_ws(wb, desired_headers)

        # Zeilen anhängen – exakt in Header-Reihenfolge
        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for info in infos:
            row_map = {
                "Zeitpunkt": now,
                "ProbenNr": info.get("ProbenNr", ""),
                "Name": info.get("Name", ""),
                "VName": info.get("Vname", ""),
                "GebDat": info.get("GebDat", ""),
                "PatID": info.get("PatID", ""),
                "Abnahme": info.get("Abnahme", ""),
                "AuftragsNr": info.get("AuftragsNr", ""),
                "Einsender": info.get("Einsender", ""),
                "Analyte": info.get("Analyte", ""),
            }
            ws.append([row_map.get(h, "") for h in headers])

        # Tabelle nur anlegen/aktualisieren, wenn mindestens eine Datenzeile existiert
        try:
            self._ensure_table(ws, headers)
        except Exception:
            pass

        try:
            wb.save(self.paths["excel_file"])
        except Exception as ex:
            print("Excel audit save failed:", ex)

    # === Delete-APIs ==========================================================
    def delete_sample_with_audit(self, proben_nr: str) -> int:
        """Einzelner Datensatz (Kompatibilität)."""
        return self.delete_samples_with_audit([proben_nr])

    def delete_samples_with_audit(self, proben_nrs: List[str]) -> int:
        """Loggt ALLE gegebenen Proben und löscht sie danach."""
        infos = []
        for pnr in proben_nrs:
            info = self.repo.get_sample_audit_info(pnr)
            if info:
                infos.append(info)
        if infos:
            self._append_audit_rows(infos)
        return self.repo.delete_samples(proben_nrs)

    # ---------------------- Singlets / Kombinationen
    def combo_stats_since(self, since: dt.datetime, top: int = 10):
        """
        Top-N für EXAKTE offene Matrizen:
          - 1er nur bei genau 1 offenem Analyt
          - 2er/3er/4er nur bei genau 2/3/4 offenen Analyten
        """
        s = since.strftime("%Y-%m-%d %H:%M:%S")
        # KEIN excluded hier – wir wollen die echte offene Matrix, nicht gefiltert
        singles, pairs, trips, quads = self.repo.open_combo_stats(s, excluded=None, max_k=4)

        def sort_desc(d: Dict) -> List[Tuple[str, int]]:
            return sorted(d.items(), key=lambda x: (-x[1], x[0]))

        srt1, srt2, srt3, srt4 = map(sort_desc, (singles, pairs, trips, quads))
        if top and top > 0:
            srt1, srt2, srt3, srt4 = srt1[:top], srt2[:top], srt3[:top], srt4[:top]
        return srt1, srt2, srt3, srt4
