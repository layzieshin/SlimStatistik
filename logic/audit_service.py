from pathlib import Path
from typing import Dict
from openpyxl import Workbook, load_workbook
from datetime import datetime

class AuditService:
    """
    Hängt gelöschte „Nicht entnommen?“-Proben an eine Excel-Datei an.
    """
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)

    def _ensure_wb(self):
        if self.excel_path.exists():
            return load_workbook(self.excel_path)
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "DeletedSamples"
        ws.append([
            "DeletedAt", "ProbenNr", "Name", "VName", "EntnahmeTag",
            "AuftragsNr", "Einsender", "Analyte"
        ])
        wb.save(self.excel_path)
        return wb

    def append_deleted(self, details: Dict[str, str]) -> None:
        wb = self._ensure_wb()
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            details.get("ProbenNr", ""),
            details.get("Name", ""),
            details.get("VName", ""),
            details.get("EntnahmeTag", ""),
            details.get("AuftragsNr", ""),
            details.get("Einsender", ""),
            details.get("AnalyteList", "")
        ])
        wb.save(self.excel_path)
