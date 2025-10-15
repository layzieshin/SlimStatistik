from typing import Iterable, Dict, List
from models.repository import Repository
import datetime, os
from openpyxl import Workbook, load_workbook

class HousekeepingService:
    def __init__(self, repo: Repository, excel_file: str):
        self.repo = repo
        self.excel_file = excel_file

    def ensure_monthly_export(self, analyte_codes: Iterable[str], today: datetime.date = None) -> str:
        """
        Führt am 1. des Monats einen Export der *ZÄHLUNG SEIT MONATSBEGINN* durch.
        Gibt den Pfad zur Excel zurück (oder existierende).
        """
        today = today or datetime.date.today()
        if today.day != 1:
            return self.excel_file  # keine Aktion heute

        # Zeitraum: aktueller Monat (1. bis heute)
        start = today.replace(day=1)
        end = today  # inkl. heute
        # aggregieren
        from logic.stats_service import StatsService
        svc = StatsService(self.repo)

        counts = {}
        for code in analyte_codes:
            n = svc.analyte_requests(code, datetime.datetime.combine(start, datetime.time.min), datetime.datetime.combine(end, datetime.time.max))
            counts[code] = n

        # Excel vorbereiten
        if os.path.exists(self.excel_file):
            wb = load_workbook(self.excel_file)
        else:
            wb = Workbook()
        sheet_name = f"{today.year}-{today.month:02d}"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in list(ws.iter_rows(min_row=2)):
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Analyt-Code", "Anzahl seit Monatsanfang", "Erstellt am"])
        for code, cnt in counts.items():
            ws.append([code, cnt, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        # Workbook speichern
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            # Default Sheet löschen, falls leer
            try:
                ws0 = wb["Sheet"]
                if ws0.max_row == 1 and ws0.max_column == 1 and ws0["A1"].value is None:
                    wb.remove(ws0)
            except Exception:
                pass
        wb.save(self.excel_file)
        return self.excel_file
